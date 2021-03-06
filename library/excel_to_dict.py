#!/usr/bin/python

from ansible.module_utils.basic import *
from openpyxl import Workbook, load_workbook
from distutils.util import strtobool
import yaml

def excel_to_dict(load_fn,yaml_fn):
    #this function excel file and transfrom it to dictonary 
    #load_fn - the excel file from were to load the data
    #yaml_fn - the destionation directory and file name to were to save the yaml
    #new_dict - is the new dictonary to be returned 
  
    wb = load_workbook(filename=load_fn)
    sheet = wb.active
    new_dict={}
    for index in range(1,sheet.max_row+1):
       new_dict[sheet.cell(row=index,column=1).value]=sheet.cell(row=index,column=2).value
    with open(yaml_fn,'w') as ymlfile:
        ymlfile.write(yaml.safe_dump(new_dict))
    return yaml.dump(new_dict)

def main():

    fields = {
       "data_file": {"default": "vlan_details.xlsx", "type": "str"},
       "yaml_file_destination": {"default": "vlan_details.yml", "type": "str"}
    }

    module = AnsibleModule(argument_spec=fields)

    
    theReturnValue = excel_to_dict(module.params["data_file"],module.params["yaml_file_destination"])
    
    module.exit_json(changed=False, meta=theReturnValue)


if __name__ == '__main__':
    main()
