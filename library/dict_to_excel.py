#!/usr/bin/python

from ansible.module_utils.basic import *
from openpyxl import Workbook, load_workbook
from distutils.util import strtobool


def data_to_excel(col, data, save_fn, newline=True):
    # get data to enter to exisiting excel file and insert it in the next free raw
    # col - column to insert the data too
    # data - data to insert to excel
    # save_fn - file to be saved
    # newline - to insert the data in a new line or in the last line
    wb = load_workbook(filename=save_fn)
    sheet = wb.active
    if newline == True:
        sheet. cell(sheet.max_row+1, col, data)
    else:
        sheet.cell(sheet.max_row, col, data)
    wb.save(filename=save_fn)


def dict_to_excel(dict_data, save_fn, newline=True):
    # get data to enter to exisiting excel file and insert it in the next free raw
    # dict_data - data to insert to excel that include column and data to be insert
    # save_fn - file to be saved
    # newline - to insert the data in a new line or in the last line
    wb = load_workbook(filename=save_fn)
    sheet = wb.active

    if newline == True:
        insert_raw = sheet.max_row+1
        for col, data in dict_data.items():
            sheet. cell(insert_raw, int(col), data)
    else:
        insert_raw = sheet.max_row
        for col, data in dict_data.items():
            sheet.cell(insert_raw, int(col), data)
    wb.save(filename=save_fn)


def main():

    fields = {
        "column": {"default": 1, "type": "int"},
        "data": {"default": " ", "type": "str"},
        "dict_data": {"default": {}, "required": False, "type": "dict"},
        "file_to_save": {"default": "test.xlsx", "type": "str"},
        "newline": {"default": True, "type": "bool"}
    }

    module = AnsibleModule(argument_spec=fields)

    if module.params["dict_data"] == {}:
        data_to_excel(module.params["column"], module.params["data"],
                      module.params["file_to_save"], module.params["newline"])
    else:
        dict_to_excel(module.params["dict_data"],
                      module.params["file_to_save"], module.params["newline"])

    theReturnValue = {"status": "data was saved"}
    module.exit_json(changed=False, meta=theReturnValue)


if __name__ == '__main__':
    main()
